import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# Step 1: Read the Excel spreadsheet
input_file = 'raw_data.xlsx'  # Replace with your input file
output_file = 'formatted_data.xlsx'

# Load the spreadsheet into a Pandas DataFrame
df = pd.read_excel(input_file, engine='openpyxl')

# Step 2: Extract and clean data
# Assuming the raw data contains hyperlinks embedded in a 'Benchmark name' column
# Extract hyperlinks if necessary (adjust column names as required)
def extract_hyperlink(cell):
    if hasattr(cell, 'hyperlink') and cell.hyperlink:
        return cell.hyperlink.target
    return cell.value

wb = load_workbook(input_file)
sheet = wb.active

# Replace 'Benchmark name' with actual column index or name
df['Benchmark name'] = [extract_hyperlink(sheet.cell(row=i+1, column=3)) for i in range(len(df))]

# Step 3: Write formatted data to a new Excel file
wb_out = load_workbook(output_file) if output_file else Workbook()
ws_out = wb_out.active

# Define column order and headers
columns = ['Framework', 'Version', 'Benchmark name', 'Provider', 'Passed assets', 'Non-compliant assets', 'Compliance posture']

# Apply column headers
ws_out.append(columns)

# Append formatted rows
for row in dataframe_to_rows(df[columns], index=False, header=False):
    ws_out.append(row)

# Apply basic formatting
for col in ws_out.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
    for cell in col:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# Save the formatted file
wb_out.save(output_file)
print(f"Formatted file saved as {output_file}")